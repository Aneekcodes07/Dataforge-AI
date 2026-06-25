"""CSV / JSON / Excel parsers producing :class:`ParsedDocument` tables.

CSV and JSON use only the standard library (no pandas) so they are lightweight
and fully unit-testable without third-party packages or network access. Excel
uses openpyxl, imported lazily.
"""

from __future__ import annotations

import csv
import io
import json
from typing import Any

from src.processing.base import ParsedDocument, ParsedTable, ProcessingError

_MAX_SNIFF_BYTES = 64 * 1024


def _decode(content: bytes) -> str:
    """Decode bytes trying common encodings, preserving as much as possible."""
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    # Last resort: never raise on decode; replace undecodable bytes.
    return content.decode("utf-8", errors="replace")


def parse_csv(content: bytes, *, name: str | None = None) -> ParsedDocument:
    """Parse delimited text (CSV/TSV) into a single table."""
    text = _decode(content)
    if not text.strip():
        raise ProcessingError("CSV file is empty")

    sample = text[:_MAX_SNIFF_BYTES]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # default to comma-delimited
    try:
        has_header = csv.Sniffer().has_header(sample)
    except csv.Error:
        has_header = True

    reader = csv.reader(io.StringIO(text), dialect)
    all_rows = [row for row in reader if row]
    if not all_rows:
        raise ProcessingError("CSV file contains no rows")

    if has_header:
        columns = [c.strip() or f"column_{i + 1}" for i, c in enumerate(all_rows[0])]
        data_rows = all_rows[1:]
    else:
        width = len(all_rows[0])
        columns = [f"column_{i + 1}" for i in range(width)]
        data_rows = all_rows

    # Normalize ragged rows to the header width.
    width = len(columns)
    normalized: list[list[Any]] = []
    for row in data_rows:
        if len(row) < width:
            row = row + [None] * (width - len(row))
        elif len(row) > width:
            row = row[:width]
        normalized.append(row)

    table = ParsedTable(columns=columns, rows=normalized, name=name)
    return ParsedDocument(
        source_type="csv",
        tables=[table],
        metadata={"delimiter": dialect.delimiter, "has_header": has_header},
    )


def _records_to_table(records: list[dict]) -> ParsedTable:
    """Convert a list of JSON objects to a table using the union of keys."""
    columns: list[str] = []
    seen: set[str] = set()
    for rec in records:
        for key in rec.keys():
            if key not in seen:
                seen.add(key)
                columns.append(str(key))
    rows: list[list[Any]] = []
    for rec in records:
        rows.append([_jsonify(rec.get(col)) for col in columns])
    return ParsedTable(columns=columns, rows=rows)


def _jsonify(value: Any) -> Any:
    """Render nested structures as compact JSON strings for tabular cells."""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def parse_json(content: bytes) -> ParsedDocument:
    """Parse JSON or JSON Lines into a table when it is record-shaped."""
    text = _decode(content).strip()
    if not text:
        raise ProcessingError("JSON file is empty")

    data: Any
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        # Try JSON Lines (one object per line).
        records = []
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                records.append(json.loads(line))
            except json.JSONDecodeError as exc:
                raise ProcessingError(f"Invalid JSON: {exc}") from exc
        data = records

    # A bare list of objects, or a dict wrapping a list of objects, becomes a table.
    if isinstance(data, list) and data and all(isinstance(r, dict) for r in data):
        return ParsedDocument(source_type="json", tables=[_records_to_table(data)])

    if isinstance(data, dict):
        for value in data.values():
            if (
                isinstance(value, list)
                and value
                and all(isinstance(r, dict) for r in value)
            ):
                return ParsedDocument(
                    source_type="json", tables=[_records_to_table(value)]
                )
        # Single object -> one-row table.
        return ParsedDocument(source_type="json", tables=[_records_to_table([data])])

    # Fallback: keep the raw JSON as text.
    return ParsedDocument(
        source_type="json", text=json.dumps(data, ensure_ascii=False, indent=2)
    )


def parse_excel(content: bytes, *, max_rows: int = 1_000_000) -> ParsedDocument:
    """Parse an .xlsx/.xls workbook; each sheet becomes a table."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - environment guard
        raise ProcessingError("openpyxl is required to parse Excel files") from exc

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - surface any openpyxl failure uniformly
        raise ProcessingError(f"Unable to read Excel workbook: {exc}") from exc

    tables: list[ParsedTable] = []
    for sheet in workbook.worksheets:
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            continue
        columns = [
            (str(c).strip() if c is not None else f"column_{i + 1}")
            for i, c in enumerate(header)
        ]
        width = len(columns)
        data_rows: list[list[Any]] = []
        for count, row in enumerate(rows_iter):
            if count >= max_rows:
                break
            values = list(row)
            if len(values) < width:
                values += [None] * (width - len(values))
            data_rows.append(values[:width])
        tables.append(ParsedTable(columns=columns, rows=data_rows, name=sheet.title))

    workbook.close()
    if not tables:
        raise ProcessingError("Excel workbook contains no readable sheets")
    return ParsedDocument(
        source_type="excel",
        tables=tables,
        metadata={"sheet_count": len(tables)},
    )
